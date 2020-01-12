from docxtpl import DocxTemplate
import docx
import openpyxl
from openpyxl import Workbook
from datetime import date


def wordInput(name, date, num1, num2):
    if num2 < 2:
        rep = 'Your color vision regarded as normal. ' \
              'You do not need a cure for color blindness.'

    elif 2 <= num2 < 4:
        rep = 'Your color vision regarded as deficient. ' \
              'Special contact lenses and glasses may help people who are color blind tell the difference between ' \
              'colors. '

    else:
        rep = 'Your color vision regarded as badly worsened. ' \
              'You can use visual aids, apps, and other technology to help you live with color blindness. ' \
              'For example, you can use an app to take a photo with your phone or tablet and then tap on part of the ' \
              'photo to find out the color of that area. '

    doc = DocxTemplate("Report.docx")
    context = {'name': name, 'date': date, 'trueAns': num1, 'falseAns': num2, 'report': rep}
    doc.render(context)
    doc.save("Report_new_" + name + '_' + date + ".docx")
    menu()





def menu():  # without arguments!!!!!!!!!!!!!
    print('Get list of games by date, enter "1"')
    print('Get list of games by user ID, enter "2"')
    print('Return, Enter "3"')
    ask = int(input('Enter here: '))
    if ask == 3:
       return
    if ask == 1:
        list1 = []
        list2 = []

        workbook = openpyxl.load_workbook("gameSQL.xlsx")
        sheet1 = workbook['Games']
        sheet2 = workbook['Users']

        ask3 = str(input('Enter your Date (for example "01.01.2020"): '))
        for i in range(2, sheet1.max_row + 1):
            if sheet1['E' + str(i)].value == ask3:
                list1.append(sheet1['A' + str(i)].value)

        print('Your games ID are: ' + str(list1))
        ask4 = int(input('Enter game ID that you want a report: '))

        for i in range(2, sheet1.max_row + 11):
            if sheet1['A' + str(i)].value == ask4:
                for j in range(2, sheet1.max_row + 1):
                    if sheet1['D' + str(i)].value == sheet2['A' + str(j)].value:
                        list2.append(sheet2['D' + str(j)].value + ' ' + sheet2['E' + str(j)].value)

        for i in range(2, sheet1.max_row + 1):
            if sheet1['A' + str(i)].value == ask4:
                list2.append(sheet1['E' + str(i)].value)
                list2.append(sheet1['B' + str(i)].value)
                list2.append(sheet1['C' + str(i)].value)

        wordInput(list2[0], list2[1], list2[2], list2[3])

    if ask == 2:

        list1 = []
        list2 = []
        ask3 = int(input('Enter your user ID: '))

        workbook = openpyxl.load_workbook("gameSQL.xlsx")
        sheet1 = workbook['Games']
        sheet2 = workbook['Users']
        for i in range(2, sheet1.max_row + 1):
            if sheet1['D' + str(i)].value == ask3:
                list1.append(sheet1['A' + str(i)].value)
        print('Your games ID are: ' + str(list1))
        ask4 = int(input('Enter game ID that you want a report: '))

        for i in range(2, sheet2.max_row + 1):
            if sheet2['A' + str(i)].value == ask3:
                list2.append(sheet2['D' + str(i)].value + ' ' + sheet2['E' + str(i)].value)

        for i in range(2, sheet1.max_row + 1):
            if sheet1['A' + str(i)].value == ask4:
                list2.append(sheet1['E' + str(i)].value)
                list2.append(sheet1['B' + str(i)].value)
                list2.append(sheet1['C' + str(i)].value)

        wordInput(list2[0], list2[1], list2[2], list2[3])
