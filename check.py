import openpyxl
from openpyxl import Workbook


def checkAnswer(answer, id, Cube):
    workbook = Workbook()
    sheet = workbook.active
    if answer == sheet[Cube + str(id)]:
        return True
    else:
        return False


def CreateGameID():
    workbook = openpyxl.load_workbook("check.xlsx")  # getting the database location
    sheet = workbook.active
    size = len(sheet['A'])
    gameid = sheet.cell(size,1).value+1
    print(sheet.cell(1, size + 1).value)
    sheet.cell(size+1, 1).value = gameid
    workbook.save("check.xlsx")


def answertoDB(answer, gameID):
    workbook = openpyxl.load_workbook("check.xlsx")  # getting the database location
    sheet = workbook.active
    temp = 2
    count = 0
    while sheet["A" + str(temp)].value != None:
        count = count + 1
        temp = int(temp) + 1

    for i in range(2, count):
        if gameID == sheet["A" + str(i)].value:
            if answer == True:
                sheet["B" + str(i)] = sheet["B" + str(i)].value + 1
                workbook.save("check.xlsx")
                break
            else:
                sheet["C" + str(i)] = sheet["C" + str(i)].value + 1
                workbook.save("check.xlsx")
                break


CreateGameID()
