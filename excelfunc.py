import openpyxl
from openpyxl import Workbook
from datetime import date

from pip._vendor.distlib.compat import raw_input


def checkAnswer(answer, card, Cube):
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Cards')
    if answer == sheet[Cube + str(card+2)].value:
        return True
    else:
        return False


def CreateGameID(user_id): #(by Tzlil) דרישה 18,22
    workbook = openpyxl.load_workbook("gameSQL.xlsx")  # getting the database location
    sheet = workbook.get_sheet_by_name('Games')
    count = 2
    while sheet['A' + str(count)].value != None:
        count = count + 1
    sheet['A' + str(count)] = count - 1
    sheet['B' + str(count)] = 0
    sheet['C' + str(count)] = 0
    sheet['D' + str(count)] = user_id
    today = date.today()
    d1 = str(today.strftime("%d.%m.%Y"))
    sheet['E' + str(count)] = d1
    workbook.save("gameSQL.xlsx")
    return count-1

def answertoDB(answer, gameID):
    workbook = openpyxl.load_workbook("gameSQL.xlsx")  # getting the database location
    sheet = workbook.get_sheet_by_name('Games')
    if answer:
        sheet["B"+str(gameID+1)].value += 1
        workbook.save("gameSQL.xlsx")

    else:
        sheet["C" + str(gameID+1)].value += 1
        workbook.save("gameSQL.xlsx")
    return True

def checkDetails(user,password):  #דרישה 1, 11,21
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Users')
    for i in range(2,sheet.max_row+1):
        if sheet['B'+str(i)].value == user:
            if sheet['C'+str(i)].value == password:
                return (True,sheet['H'+str(i)].value,sheet['A'+str(i)].value)
            else:
                return (False,'badpassword')
    return (False,'badusername')


def addUser(username, password, name, lastname, email, phone, type):  # דרישה: 12,3
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Users')
    count = 2
    list = []
    for i in range(2, sheet.max_row + 1):
        if sheet['B' + str(i)].value != None:
            list.append(sheet['B' + str(i)].value)
    for i in list:
        if username == i:
            return False
    while sheet['A' + str(count)].value != None:
        count = count + 1
    sheet['A' + str(count)] = count - 1
    sheet['B' + str(count)] = username
    sheet['C' + str(count)] = password
    sheet['D' + str(count)] = name
    sheet['E' + str(count)] = lastname
    sheet['F' + str(count)] = email
    sheet['G' + str(count)] = phone
    sheet['H' + str(count)] = type
    workbook.save("gameSQL.xlsx")
    return True

def editUser(i_d,edit,edit_choice):
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Users')
    for i in range(2, sheet.max_row + 1):
        if sheet['A' + str(i)].value == i_d:
            if edit_choice == 'Username':
                sheet['B' + str(i)] = edit
            if edit_choice == 'Password':
                sheet['C' + str(i)] = edit
            if edit_choice == 'Name':
                sheet['D' + str(i)] = edit
            if edit_choice == 'Lastname':
                sheet['E' + str(i)] = edit
            if edit_choice == 'email':
                sheet['F' + str(i)] = edit
            if edit_choice == 'phone':
                sheet['G' + str(i)] = edit
            if edit_choice == 'type':
                sheet['H' + str(i)] = edit
    workbook.save("gameSQL.xlsx")
    return True

def deleteUser(i_d):  # דרישה 7,15
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Users')
    for i in range(2, sheet.max_row + 1):
        if sheet['A' + str(i)].value == i_d:
            sheet['A' + str(i)] = None
            sheet['B' + str(i)] = None
            sheet['C' + str(i)] = None
            sheet['D' + str(i)] = None
            sheet['E' + str(i)] = None
            sheet['F' + str(i)] = None
            sheet['G' + str(i)] = None
            sheet['H' + str(i)] = None
            workbook.save("gameSQL.xlsx")
            workbook = openpyxl.load_workbook("gameSQL.xlsx")
            sheet1 = workbook.get_sheet_by_name('Games')
            for j in range(2, sheet1.max_row + 1):
                if sheet1['D' + str(j)].value == i_d:
                    sheet1['A' + str(i)] = None
                    sheet1['B' + str(i)] = None
                    sheet1['C' + str(i)] = None
                    sheet1['D' + str(i)] = None
                    sheet1['E' + str(i)] = None
                    sheet1['F' + str(i)] = None
                    workbook.save("gameSQL.xlsx")
            return True
    return False


def resetUser(i_d):  # דרישה 16
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Games')
    for i in range(2, sheet.max_row + 1):
        if sheet['D' + str(i)].value == i_d:
            sheet['A' + str(i)] = None
            sheet['B' + str(i)] = None
            sheet['C' + str(i)] = None
            sheet['D' + str(i)] = None
            sheet['E' + str(i)] = None
            sheet['F' + str(i)] = None


            workbook.save("gameSQL.xlsx")
    return True


def gameOfDate(day, month, year):  # דרישה 17
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Games')
    date = str(day) + '.' + str(month) + '.' + year
    list = []
    for i in range(2, sheet.max_row + 1):
        if str(sheet['E' + str(i)].value) == date:
            list.append(sheet['A' + str(i)].value)
    return list


def gamesOfUser(usename):  # דרישה 13,5
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Users')
    list = []
    for i in range(2, sheet.max_row + 1):
        if str(sheet['B' + str(i)].value) == usename:
            i_d = sheet['A' + str(i)].value
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet1 = workbook.get_sheet_by_name('Games')
    for i in range(2, sheet1.max_row + 1):
        if sheet1['D' + str(i)].value == i_d:
            list.append(sheet1['A' + str(i)].value)
    return list



def Card_enable(card_num): #דרישה 8.2
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Cards')
    for i in range(2, sheet.max_row + 1):
        if sheet['A' + str(i)].value == card_num:
            if str(sheet['I' + str(i)].value) == 'yes':
                sheet['I' + str(i)].value = 'no'
                workbook.save("gameSQL.xlsx")
                return None
            if str(sheet['I' + str(i)].value) == 'no':
                sheet['I' + str(i)].value = 'yes'
                workbook.save("gameSQL.xlsx")
                return None

def Print_User_list(): #דרישה 6
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Users')
    for i in range(2, sheet.max_row + 1):
        print("ID: {}  Name: {}  Lastname: {}".format(str(sheet['A' + str(i)].value),str(sheet['D' + str(i)].value),str(sheet['E' + str(i)].value)))


def feedback(game_id):
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Games')
    ask = str(input('do you want to give feedback?(yes or no) '))
    if ask== "yes":
     feed = raw_input("Please enter your feedback on the game")
     sheet['F'+str(game_id+1)].value = feed
     workbook.save("gameSQL.xlsx")
    if ask== "no":
        feed = "אין"
        sheet['F' + str(game_id + 1)].value = feed
        workbook.save("gameSQL.xlsx")
    print("Thank you, you can exit the game now")

def Edit_Report():
    workbook = openpyxl.load_workbook("gameSQL.xlsx")
    sheet = workbook.get_sheet_by_name('Report')
    choice = int(input("choose msg report number to change:"))
    rep = raw_input("Enter new report msg:")
    sheet['B'+str(choice+1)].value = rep
    workbook.save("gameSQL.xlsx")

